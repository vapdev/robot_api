from openpyxl.styles import Border, Side 
from openpyxl.styles import Alignment
import pandas as pd
import requests

def classifica(sol_pai, sols_filhas, db):
    _sols_que_ja_foram = []
    
    for x in sols_filhas:
        query_sol = f"select tp_solicitacao, nr_solicitacao, ds_assunto, st_solic, ds_status from solicservico_v2 a, statussolic b where a.nr_solicitacao = {x} and a.st_solic = b.cd_status"
        query = db.query(query_sol)
        
        # atribuo as variaveis que vieram da query do banco
        nr_solicitacao = query['nr_solicitacao'].values[0]
        ds_assunto = query['ds_assunto'].values[0]
        ds_status = query['ds_status'].values[0]
        tp_solicitacao = query['tp_solicitacao'].values[0]
        horas_realizadas = formata_horas(db.query(f'SELECT retorna_min_realizados_solic({nr_solicitacao})/60 minutos FROM DUAL')['minutos'].values[0])
        
        
        query_gaps = f"select PACK_ORCAMENTO.tem_orcamento_rentavel_arvore({nr_solicitacao}) from dual"
        if db.query(query_gaps).values[0] == 'S':
            _sols_que_ja_foram.append(nr_solicitacao)
        
        if ds_assunto.find('Deslocamento') > 0 or ds_assunto.find('deslocamento') > 0 or ds_assunto.find('DESLOCAMENTO') > 0 and tp_solicitacao not in [2,3,10,12]:
            _sols_que_ja_foram.append(nr_solicitacao)
        
    return _sols_que_ja_foram

def aplica_borda_total():
    return Border(top = Side(border_style='thin', color='8DB4E2'),    
                                             right = Side(border_style='thin', color='8DB4E2'), 
                                             bottom = Side(border_style='thin', color='8DB4E2'),
                                             left = Side(border_style='thin', color='8DB4E2'))

def aplica_borda_sem_esquerda():
    return Border(top = Side(border_style='thin', color='8DB4E2'),    
                                             right = Side(border_style='thin', color='8DB4E2'), 
                                             bottom = Side(border_style='thin', color='8DB4E2'))

def aplica_borda_sem_direita():
    return Border(top = Side(border_style='thin', color='8DB4E2'),    
                                             left = Side(border_style='thin', color='8DB4E2'), 
                                             bottom = Side(border_style='thin', color='8DB4E2'))
    
def formata_horas(horas_realizadas):
    if horas_realizadas:
        horas_realizadas = horas_realizadas.item()
    seconds = horas_realizadas*60*60
    minutes, seconds = divmod(seconds, 60)
    hours, minutes = divmod(minutes, 60)
    horas_realizadas = "%02d:%02d:%02d"%(hours,minutes,seconds)
    return horas_realizadas

def faz_total(count, start, horas_total, ws):
    last_sol = count + start - 1
    lista = str(ws.merged_cells).split(' ')
    e_merged = f'B{last_sol+1}:F{last_sol+1}'
    if e_merged in lista:
        ws.unmerge_cells(e_merged)
    print('last sol '+ str(last_sol) + '-- count ' + str(count))
    ws[f'F{last_sol+1}'] = f'=SUM(F{start}:F{last_sol})'
    ws[f'G{last_sol+1}'].number_format = '[h]:mm:ss'
    ws[f'G{last_sol+1}'] = f'=SUM(G{start}:G{last_sol})'
    ws[f'G{last_sol+1}'] = formata_horas(horas_total)
    ws[f'F{last_sol+1}'].alignment = Alignment(horizontal='center')
    ws[f'G{last_sol+1}'].alignment = Alignment(horizontal='center')
    ws[f'G{last_sol+1}'].border = aplica_borda_total()
    ws[f'C{last_sol+1}'].border = aplica_borda_total()
    ws[f'D{last_sol+1}'].border = aplica_borda_total()
    ws[f'E{last_sol+1}'].border = aplica_borda_total()
    ws[f'F{last_sol+1}'].border = aplica_borda_total()
    ws[f'G{last_sol+1}'].border = aplica_borda_total()
    return ws[f'G{last_sol+1}']

def preenche_linha(valor, coluna, linha, alinhamento, borda, ws, planilha):
    if valor:
        ws[f'{coluna}{linha}'] = valor 
    ws[f'{coluna}{linha}'].alignment = Alignment(horizontal=f'{alinhamento}')
    if borda == 'no_right':
        ws[f'{coluna}{linha}'].border = aplica_borda_sem_direita()
    if borda == 'no_left':
        ws[f'{coluna}{linha}'].border = aplica_borda_sem_esquerda()
    if borda == 'total':
        ws[f'{coluna}{linha}'].border = aplica_borda_total()
    planilha.save('teste.xlsx')
    
def get_proj_id(nr_solic):
    url = f"https://maxicon.teamwork.com/search.json?searchFor=projects&searchTerm={nr_solic}"
    request = requests.get(url, headers={'Authorization':'Basic dHdwX0g5ZEtmbTBqckdPM2lYWWEzWTFreXg5a0ZzVE06eHh4'}) 
    projetos = request.json()
    project_id = projetos['searchResult']['projects'][0]['id']
    return project_id


def acha_start(tipo, ws):
    if tipo == 'atendimento':
        for row in ws.iter_rows(2):
            for cell in row:
                if cell.value == "ATENDIMENTO REMOTO E PRESENCIAL":
                    atendimento_start = cell.row + 2
                    return atendimento_start
    
    if tipo == 'deslocamentos':
        for row in ws.iter_rows(2):
            for cell in row:
                if cell.value == "DESLOCAMENTOS -Possuem Tratamento Comercial Distinto":
                    deslocamentos_start = cell.row + 2
                    return deslocamentos_start
                    
    if tipo == 'gaps':
        for row in ws.iter_rows(2):
            for cell in row:
                if cell.value == "GAPS/MELHORIAS-Possuem Tratamento Comercial Distinto":
                    gaps_start = cell.row + 2 
                    return gaps_start
                    
    if tipo == 'ajustes':                    
        for row in ws.iter_rows(2):
            for cell in row:
                if cell.value == "AJUSTES DE PROCESSOS - Não contam como Hora de Projeto":
                    ajustes_start = cell.row + 2 
                    return ajustes_start
                    
    if tipo == 'suporte':                    
        for row in ws.iter_rows(2):
            for cell in row:
                if cell.value == "SOLICITAÇÕES DE SUPORTE  - Não contam como Hora de Projeto":
                    suporte_start = cell.row + 2
                    return suporte_start 
                


